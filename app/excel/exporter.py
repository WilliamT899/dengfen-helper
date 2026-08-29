"""Excel 导出：多 Sheet（按班级）+ 统计公式块。

布局（每班一个 Sheet，Sheet 名=班级名）：
  A=序号 B=学号 C=姓名 D=分数（第 2 行起，未拍学生留空）
  统计块：F 列标签，G 列公式（右侧独立区域，全部用公式，改分自动重算）
总览 Sheet：全部班级合并展示。
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.roster import Roster, Student, _id_sort_key

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(*[Side(style="thin", color="B0B0B0")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")
STAT_FILL = PatternFill("solid", fgColor="FFF2CC")

STAT_ITEMS: Sequence[Tuple[str, str]] = (
    ("平均分", '=IFERROR(AVERAGE({rng}),"")'),
    ("中位数", '=IFERROR(MEDIAN({rng}),"")'),
    ("最高分", '=IFERROR(MAX({rng}),"")'),
    ("最低分", '=IFERROR(MIN({rng}),"")'),
    ("100分人数", '=COUNTIF({rng},100)'),
    ("95分以上人数", '=COUNTIF({rng},">=95")'),
    ("90分以上人数", '=COUNTIF({rng},">=90")'),
    ("80分以上人数", '=COUNTIF({rng},">=80")'),
    ("60分以下人数", '=COUNTIFS({rng},"<60",{rng},"<>")'),
)


def export_excel(
    roster: Roster,
    scores: Dict[Tuple[str, str], Optional[float]],  # (学号, 姓名) -> 分数；None=未拍
    path: Path,
) -> Path:
    """导出成绩表。scores 键为 (学号, 姓名) 以兼容同名不同班场景。"""
    wb = Workbook()
    wb.remove(wb.active)

    for klass, students in roster.by_class().items():
        _write_class_sheet(wb, klass, students, scores)

    _write_overview_sheet(wb, roster, scores)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def _header_row(ws) -> None:
    for col, title in enumerate(("序号", "学号", "姓名", "分数"), start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _write_class_sheet(wb: Workbook, klass: str, students: List[Student],
                       scores: Dict[Tuple[str, str], Optional[float]]) -> None:
    ws = wb.create_sheet(title=_safe_sheet_name(klass))
    _header_row(ws)

    for i, st in enumerate(students):
        row = i + 2
        score = scores.get((st.student_id, st.name))
        ws.cell(row=row, column=1, value=i + 1).alignment = CENTER
        ws.cell(row=row, column=2, value=st.student_id).alignment = CENTER  # 文本写入，保留前导0
        ws.cell(row=row, column=3, value=st.name).alignment = CENTER
        sc = ws.cell(row=row, column=4, value=score if score is not None else None)
        sc.alignment = CENTER
        sc.number_format = "0.00"

    n = len(students)
    last = 1 + n
    rng = f"$D$2:$D${last}"

    # 统计块：F 列标签，G 列公式
    ws.cell(row=1, column=6, value="统计项").font = HEADER_FONT
    ws.cell(row=1, column=7, value="结果").font = HEADER_FONT
    ws.cell(row=1, column=6).fill = STAT_FILL
    ws.cell(row=1, column=7).fill = STAT_FILL
    for i, (label, formula) in enumerate(STAT_ITEMS):
        lc = ws.cell(row=2 + i, column=6, value=label)
        lc.fill = STAT_FILL
        fc = ws.cell(row=2 + i, column=7, value=formula.format(rng=rng))
        fc.fill = STAT_FILL
        fc.number_format = "0.00" if i < 4 else "0"

    for col, width in ((1, 6), (2, 10), (3, 12), (4, 10), (6, 14), (7, 10)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_overview_sheet(wb: Workbook, roster: Roster,
                          scores: Dict[Tuple[str, str], Optional[float]]) -> None:
    ws = wb.create_sheet(title="总览")
    for col, title in enumerate(("班级", "学号", "姓名", "分数"), start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.freeze_panes = "A2"

    rows = sorted(roster.students, key=lambda s: (s.klass, _id_sort_key(s)))
    for i, st in enumerate(rows):
        row = i + 2
        score = scores.get((st.student_id, st.name))
        ws.cell(row=row, column=1, value=st.klass).alignment = CENTER
        ws.cell(row=row, column=2, value=st.student_id).alignment = CENTER
        ws.cell(row=row, column=3, value=st.name).alignment = CENTER
        sc = ws.cell(row=row, column=4, value=score if score is not None else None)
        sc.alignment = CENTER
        sc.number_format = "0.00"

    for col, width in ((1, 12), (2, 10), (3, 12), (4, 10)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _safe_sheet_name(name: str) -> str:
    """Sheet 名清洗：去掉 Excel 非法字符，限长 31。"""
    for ch in r"[]:*?/\\":
        name = name.replace(ch, "")
    return (name or "未分班")[:31]
