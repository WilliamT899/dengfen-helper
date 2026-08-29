"""学生名单导入与索引。

支持 .xlsx（openpyxl）与 .csv；自动在前 5 行内查找表头行，识别
"姓名/学号/班级"列（兼容常见别名），返回结构化名单。
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

NAME_ALIASES = ("姓名", "学生姓名", "名字", "学生")
ID_ALIASES = ("学号", "学生学号", "编号", "学籍号")
CLASS_ALIASES = ("班级", "班", "班别", "行政班")

HEADER_SCAN_ROWS = 5  # 在前 5 行内找表头


@dataclass
class Student:
    name: str
    student_id: str   # 规范化字符串，保留前导 0
    klass: str        # 班级名；名单无班级列时为空


@dataclass
class Roster:
    students: List[Student]
    source: str = ""

    def __len__(self) -> int:
        return len(self.students)

    def classes(self) -> List[str]:
        seen: List[str] = []
        for s in self.students:
            if s.klass not in seen:
                seen.append(s.klass)
        return seen

    def by_class(self) -> Dict[str, List[Student]]:
        groups: Dict[str, List[Student]] = {}
        for s in self.students:
            groups.setdefault(s.klass, []).append(s)
        for lst in groups.values():
            lst.sort(key=_id_sort_key)
        return groups

    def id_index(self) -> Dict[str, List[Student]]:
        idx: Dict[str, List[Student]] = {}
        for s in self.students:
            if s.student_id:
                idx.setdefault(s.student_id, []).append(s)
        return idx


def _id_sort_key(s: Student):
    # 纯数字按数值排，其余按字符串排
    return (0, int(s.student_id)) if s.student_id.isdigit() else (1, s.student_id)


def _norm_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _find_columns(header_row: Tuple[str, ...]) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """在表头行中匹配 姓名/学号/班级 列下标。"""
    name_col = id_col = class_col = None
    for idx, cell in enumerate(header_row):
        val = _norm_cell(cell)
        if name_col is None and val in NAME_ALIASES:
            name_col = idx
        elif id_col is None and val in ID_ALIASES:
            id_col = idx
        elif class_col is None and val in CLASS_ALIASES:
            class_col = idx
    return name_col, id_col, class_col


def load_roster(path: Path) -> Roster:
    """读取名单 Excel。找不到表头列时抛 ValueError（UI 提示用户手动指定）。"""
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 找表头行
    header_idx = -1
    name_col = id_col = class_col = None
    for i, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        cols = _find_columns(tuple(row))
        if any(c is not None for c in cols):
            header_idx, (name_col, id_col, class_col) = i, cols
            break
    if header_idx < 0:
        raise ValueError("未在文件中找到'姓名/学号/班级'表头，请检查名单格式（首行应为表头）")
    if name_col is None and id_col is None:
        raise ValueError("表头中缺少'姓名'和'学号'列，请检查名单格式")

    students: List[Student] = []
    for row in rows[header_idx + 1:]:
        name = _norm_cell(row[name_col]) if name_col is not None and name_col < len(row) else ""
        sid = _norm_cell(row[id_col]) if id_col is not None and id_col < len(row) else ""
        klass = _norm_cell(row[class_col]) if class_col is not None and class_col < len(row) else ""
        if not name and not sid:
            continue  # 空行
        students.append(Student(name=name or f"(无姓名-{sid or len(students)+1})",
                                student_id=sid, klass=klass or "未分班"))

    if not students:
        raise ValueError("名单中没有学生数据")
    return Roster(students=students, source=str(path))
