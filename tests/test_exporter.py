"""Excel 导出测试：多 Sheet、统计公式、留空处理。"""
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.excel.exporter import export_excel
from app.roster import Roster, Student


@pytest.fixture
def roster():
    students = [
        Student("邢子洋", "01", "三年级1班"),
        Student("陈曦", "02", "三年级1班"),
        Student("张梓凑", "03", "三年级1班"),   # 未拍：留空
        Student("朱芷瑶", "01", "三年级2班"),
    ]
    return Roster(students=students)


def test_export_structure(roster, tmp_path):
    path = tmp_path / "成绩.xlsx"
    scores = {("01", "邢子洋"): 98.5, ("02", "陈曦"): 87, ("01", "朱芷瑶"): 100}
    export_excel(roster, scores, path)

    wb = load_workbook(str(path))
    assert set(wb.sheetnames) == {"三年级1班", "三年级2班", "总览"}

    ws = wb["三年级1班"]
    assert ws["A1"].value == "序号"
    assert ws["B2"].value == "01"          # 学号文本，保留前导0
    assert ws["D2"].value == 98.5
    assert ws["D4"].value is None          # 未拍留空
    # 统计公式
    assert ws["G2"].value == '=IFERROR(AVERAGE($D$2:$D$4),"")'
    assert ws["G6"].value == "=COUNTIF($D$2:$D$4,100)"
    assert ws["G10"].value == '=COUNTIFS($D$2:$D$4,"<60",$D$2:$D$4,"<>")'

    ws2 = wb["总览"]
    assert ws2["A2"].value == "三年级1班"
    assert ws2["B2"].value == "01"


def test_sheet_name_sanitize(roster, tmp_path):
    roster.students[0].klass = '三[1]班:测试'
    path = tmp_path / "x.xlsx"
    export_excel(roster, {}, path)
    wb = load_workbook(str(path))
    assert "三1班测试" in wb.sheetnames
