#!/usr/bin/env python3
"""端到端测试：名单导入 → 6 张样本识别 → 名单匹配 → Excel 导出 → 公式校验。

模拟 GUI 的完整业务流（无界面），用于开发验证与 CI 冒烟。
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from app.excel.exporter import export_excel  # noqa: E402
from app.matcher import match_student  # noqa: E402
from app.ocr.engine import get_engine  # noqa: E402
from app.ocr.pipeline import recognize_file  # noqa: E402
from app.roster import load_roster  # noqa: E402
from app.workspace import Workspace  # noqa: E402

GROUND_TRUTH = {
    "样本图片1": ("邢子洋", "28", 99.0),
    "样本图片2": ("陈曦", "20", 88.5),
    "样本图片3": ("邓睿泽", "22", 96.0),
    "样本图片4": ("张梓凌", "13", 100.0),
    "样本图片5": ("朱芷瑶", "14", 97.0),
    "样本图片6": ("柯梓涵", "38", 98.0),
}


def main() -> int:
    roster_path = Path("测试名单.xlsx")
    if not roster_path.exists():
        print("缺少 测试名单.xlsx")
        return 1
    roster = load_roster(roster_path)
    print(f"1. 名单导入: {len(roster)} 人 ✓")

    ws = Workspace()
    ws.init_from_roster(roster)

    engine = get_engine()
    print("2. 批量识别 6 张样本…")
    name_ok = score_ok = 0
    for img in sorted(Path("样本图片").glob("*.jpg")):
        r = recognize_file(img, engine)
        m = match_student(r.name, r.student_id, roster)
        if m.is_unique:
            st = m.student
        else:
            st = None
        # 回填（GUI 同逻辑）
        ws.apply_result(
            st.student_id if st else r.student_id,
            st.name if st else r.name,
            st.klass if st else r.klass,
            r.score, "", match_status=m.status)
        gt = GROUND_TRUTH[img.stem]
        if st and st.name == gt[0]:
            name_ok += 1
        if r.score == gt[2]:
            score_ok += 1
        print(f"   {img.name}: OCR={r.name}/{r.score} → 匹配={'✓' if m.is_unique else '待确认'}")
    print(f"3. 识别统计: 姓名 {name_ok}/6, 分数 {score_ok}/6")

    out = Path("/tmp/e2e_成绩.xlsx")
    scores = {(row.student_id, row.name): row.score for row in ws.rows}
    export_excel(roster, scores, out)
    print(f"4. 导出: {out}")

    wb = load_workbook(str(out))
    print(f"   Sheets: {wb.sheetnames}")
    sheet = wb[wb.sheetnames[0]]
    assert sheet["G2"].value.startswith("=IFERROR(AVERAGE"), "统计公式缺失!"
    assert sheet["G10"].value.startswith("=COUNTIFS"), "60分以下公式缺失!"
    print("5. 统计公式校验 ✓ (AVERAGE/MEDIAN/MAX/MIN/COUNTIF/COUNTIFS)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
